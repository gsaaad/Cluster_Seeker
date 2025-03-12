import os
import pandas as pd
import glob
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def combine_csv_files(input_dir, output_csv, output_excel):
    all_files = glob.glob(os.path.join(input_dir, "*.csv"))

    if not all_files:
        print(f"No CSV files found in {input_dir}")
        return None

    df_list = []
    for filename in tqdm(all_files, desc="Combining CSV files"):
        df = pd.read_csv(filename)
        df_list.append(df)

    combined_df = pd.concat(df_list, ignore_index=True)

    # Save as CSV
    combined_df.to_csv(output_csv, index=False)
    print(f"Combined CSV file saved as: {output_csv}")

    # Save as Excel with sheets for each extension
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Add 'All Files' sheet
    ws = wb.create_sheet("All Files")
    for r in dataframe_to_rows(combined_df, index=False, header=True):
        ws.append(r)

    # Add sheets for each extension
    for ext, group in combined_df.groupby('File Extension'):
        sheet_name = ext.strip('.') if ext else 'No Extension'
        sheet_name = sheet_name[:31]  # Excel sheet names have a 31 character limit
        ws = wb.create_sheet(sheet_name)
        for r in dataframe_to_rows(group, index=False, header=True):
            ws.append(r)

    wb.save(output_excel)
    print(f"Combined Excel file saved as: {output_excel}")

    return combined_df

def find_duplicates(df):
    # Group by 'File Name' and 'File Size'
    duplicates = df[df.duplicated(['File Name', 'File Size'], keep=False)]
    duplicates = duplicates.sort_values(['File Name', 'File Size'])
    return duplicates

def save_duplicates(duplicates_df, output_csv, output_excel):
    # Save as CSV
    duplicates_df.to_csv(output_csv, index=False)
    print(f"Duplicates CSV file saved as: {output_csv}")

    # Save as Excel with sheets for each extension
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Add 'All Duplicates' sheet
    ws = wb.create_sheet("All Duplicates")
    for r in dataframe_to_rows(duplicates_df, index=False, header=True):
        ws.append(r)

    # Add sheets for each extension
    for ext, group in duplicates_df.groupby('File Extension'):
        sheet_name = ext.strip('.') if ext else 'No Extension'
        sheet_name = sheet_name[:31]  # Excel sheet names have a 31 character limit
        ws = wb.create_sheet(sheet_name)
        for r in dataframe_to_rows(group, index=False, header=True):
            ws.append(r)

    wb.save(output_excel)
    print(f"Duplicates Excel file saved as: {output_excel}")

if __name__ == '__main__':
    input_dir = "output/file_batches"
    combined_csv = "combined_files.csv"
    combined_excel = "combined_files.xlsx"
    duplicates_csv = "duplicates.csv"
    duplicates_excel = "duplicates.xlsx"

    # Combine all CSV files
    combined_df = combine_csv_files(input_dir, combined_csv, combined_excel)

    if combined_df is not None:
        # Find duplicates
        print("Finding duplicates...")
        duplicates_df = find_duplicates(combined_df)

        if duplicates_df.empty:
            print("No duplicates found.")
        else:
            # Save duplicates to CSV and Excel files
            save_duplicates(duplicates_df, duplicates_csv, duplicates_excel)
            print(f"Number of duplicates found: {len(duplicates_df)}")

        # Print some statistics
        print("\nStatistics:")
        print(f"Total files processed: {len(combined_df)}")
        print(f"Unique file names: {combined_df['File Name'].nunique()}")
        print(f"Unique file paths: {combined_df['File Path'].nunique()}")
